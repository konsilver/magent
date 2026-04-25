"""Unit tests for core.content.artifact_summary.

Each test feeds bytes of a specific type and checks the summary shape.
These are pure-logic tests — no DB, no storage.
"""

import io

import pytest

from core.content.artifact_summary import build_summary


# ── Plain text & markdown ─────────────────────────────────────────────────────

def test_text_document_truncates_to_500_chars():
    body = ("hello world " * 200).encode("utf-8")  # > 500 chars
    summary = build_summary(body, "readme.txt", "text/plain")
    # First 500 chars + "（共 X 字）"
    assert "共" in summary and "字" in summary
    assert summary.startswith("hello world")
    assert len(summary) <= 1000  # hard cap


def test_text_document_short_shows_full():
    body = "short content".encode("utf-8")
    summary = build_summary(body, "note.txt", "text/plain")
    assert "short content" in summary
    assert "共" in summary


def test_markdown_keeps_first_30_lines():
    body = "\n".join(f"# Heading {i}" for i in range(50)).encode("utf-8")
    summary = build_summary(body, "notes.md", "text/markdown")
    assert "# Heading 0" in summary
    assert "# Heading 29" in summary
    # Line 30+ should NOT be in summary
    assert "# Heading 30" not in summary
    assert "共 50 行" in summary


# ── Code ──────────────────────────────────────────────────────────────────────

def test_code_py_truncates():
    lines = [f"def f{i}(): pass" for i in range(100)]
    body = "\n".join(lines).encode("utf-8")
    summary = build_summary(body, "module.py", "text/x-python")
    assert "def f0()" in summary
    assert "def f29()" in summary
    assert "def f30()" not in summary
    assert "共 100 行" in summary


# ── XLSX ──────────────────────────────────────────────────────────────────────

def test_xlsx_shows_sheets_headers_rows():
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "员工信息"
    ws1.append(["姓名", "部门", "岗位"])
    ws1.append(["张三", "工程部", "工程师"])
    ws1.append(["李四", "市场部", "经理"])

    ws2 = wb.create_sheet("薪酬数据")
    ws2.append(["工号", "基本薪资", "奖金"])
    ws2.append(["E001", 10000, 5000])

    buf = io.BytesIO()
    wb.save(buf)

    summary = build_summary(buf.getvalue(), "hr.xlsx",
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    assert "共 2 个 sheet" in summary
    assert "员工信息" in summary
    assert "薪酬数据" in summary
    assert "姓名" in summary
    assert "张三" in summary


def test_xlsx_empty_sheet():
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    wb = Workbook()
    wb.active.title = "Sheet1"
    buf = io.BytesIO()
    wb.save(buf)

    summary = build_summary(buf.getvalue(), "empty.xlsx", "")
    assert "共 1 个 sheet" in summary
    assert "Sheet1" in summary


# ── CSV ───────────────────────────────────────────────────────────────────────

def test_csv_shows_header_and_rows():
    body = "姓名,部门,岗位\n张三,工程部,工程师\n李四,市场部,经理\n王五,HR,专员\n".encode("utf-8")
    summary = build_summary(body, "staff.csv", "text/csv")
    assert "CSV" in summary
    assert "姓名 | 部门 | 岗位" in summary
    assert "张三" in summary


def test_csv_empty():
    summary = build_summary(b"", "empty.csv", "text/csv")
    assert "空" in summary


# ── Image ─────────────────────────────────────────────────────────────────────

def test_image_shows_mime_and_size():
    # Fake image bytes — PIL will fail, we should still get a basic summary
    body = b"\x89PNG\r\n\x1a\n" + b"\x00" * 100
    summary = build_summary(body, "logo.png", "image/png")
    assert "[图片]" in summary
    assert "logo.png" in summary
    assert "image/png" in summary


# ── Generic fallback ──────────────────────────────────────────────────────────

def test_unknown_extension_falls_back_to_generic():
    body = b"binary garbage" * 100
    summary = build_summary(body, "unknown.zip", "application/zip")
    assert "unknown.zip" in summary
    assert "application/zip" in summary


# ── Error resilience ─────────────────────────────────────────────────────────

def test_corrupted_xlsx_falls_back_to_generic_or_parse():
    # Not a real xlsx — openpyxl will fail
    body = b"not an xlsx"
    summary = build_summary(body, "bad.xlsx", "")
    # Must not raise — should return something
    assert isinstance(summary, str)
    assert len(summary) > 0


def test_build_summary_never_raises():
    # Garbage bytes across all extension paths
    for filename in ["bad.pdf", "bad.docx", "bad.xlsx", "bad.csv",
                     "bad.pptx", "bad.md", "bad.py", "unknown.xyz"]:
        summary = build_summary(b"\x00\x01\x02", filename, "")
        assert isinstance(summary, str)


# ── Summary hard cap ─────────────────────────────────────────────────────────

# ── build_summary_from_text (the frontend-already-parsed path) ────────────

def test_from_text_long_pdf_like_truncates_500():
    from core.content.artifact_summary import build_summary_from_text
    text = ("本文档为年度财报，涵盖 " * 200)
    out = build_summary_from_text(text, "report.pdf", "application/pdf")
    assert "共" in out and "字" in out
    assert len(out) <= 1001


def test_from_text_short_text_returned_directly():
    from core.content.artifact_summary import build_summary_from_text
    out = build_summary_from_text("hello", "note.txt", "text/plain")
    # short strings return full text; cap 600 chars triggers this path
    assert "hello" in out


def test_from_text_markdown_uses_line_truncation():
    from core.content.artifact_summary import build_summary_from_text
    text = "\n".join(f"# H{i}" for i in range(50))
    out = build_summary_from_text(text, "notes.md", "text/markdown")
    assert "# H0" in out
    assert "# H29" in out
    assert "# H30" not in out
    assert "共 50 行" in out


def test_from_text_code_uses_line_truncation():
    from core.content.artifact_summary import build_summary_from_text
    text = "\n".join(f"def f{i}(): pass" for i in range(100))
    out = build_summary_from_text(text, "m.py", "text/x-python")
    assert "def f29" in out
    assert "def f30" not in out
    assert "共 100 行" in out


def test_from_text_csv_shows_first_rows():
    from core.content.artifact_summary import build_summary_from_text
    text = "\n".join([f"r{i},c{i},d{i}" for i in range(20)])
    out = build_summary_from_text(text, "data.csv", "text/csv")
    assert "r0" in out
    assert "r5" in out
    assert "r6" not in out  # only first 6 lines
    assert "共 20 行" in out


def test_from_text_xlsx_preserves_sheet_markers():
    """When the parsed markdown uses `### Sheet: N` headers, the truncation
    should preferentially cut at sheet boundaries."""
    from core.content.artifact_summary import build_summary_from_text
    text = (
        "### Sheet 1: 员工\n| 姓名 | 部门 |\n|---|---|\n| 张三 | 工程 |\n" * 20
        + "### Sheet 2: 薪酬\n| 工号 | 金额 |\n|---|---|\n| E001 | 10000 |\n" * 20
    )
    out = build_summary_from_text(
        text, "book.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    assert "### Sheet" in out
    assert len(out) <= 1001


def test_from_text_empty_returns_metadata_only():
    from core.content.artifact_summary import build_summary_from_text
    out = build_summary_from_text("", "mystery.bin", "application/octet-stream")
    assert "mystery.bin" in out
    assert "octet-stream" in out


def test_from_text_whitespace_only_returns_metadata_only():
    from core.content.artifact_summary import build_summary_from_text
    out = build_summary_from_text("   \n  \t  ", "blank.txt", "text/plain")
    assert "blank.txt" in out


def test_summary_respects_hard_cap():
    pytest.importorskip("openpyxl")
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Huge"
    # Generate many rows with long cell values
    for i in range(100):
        ws.append([f"col{j}_value_is_long_enough_{i}" for j in range(20)])
    buf = io.BytesIO()
    wb.save(buf)

    summary = build_summary(buf.getvalue(), "huge.xlsx", "")
    assert len(summary) <= 1001  # _MAX_SUMMARY_CHARS + "…"
